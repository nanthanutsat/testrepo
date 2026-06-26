"""
Export workbook definition to Excel file.
"""
from pathlib import Path
from openpyxl import Workbook
from workbook_definition import service_request_workbook, formulas


def export_workbook_to_excel(workbook, formula_lib, output_path: Path) -> None:
    """
    Export Python workbook definition to an Excel file.
    
    Args:
        workbook: Workbook instance
        formula_lib: FormulaLibrary instance
        output_path: Path to output .xlsx file
    """
    wb = Workbook()
    first = True

    for worksheet in workbook.worksheets:
        if first:
            ws = wb.active
            ws.title = worksheet.name
            first = False
        else:
            ws = wb.create_sheet(title=worksheet.name)

        for table in worksheet.tables:
            # Write headers
            for col_idx, column in enumerate(table.columns, start=1):
                ws.cell(row=1, column=col_idx, value=column.header)

            # Write rows
            for row_idx, row in enumerate(table.rows, start=2):
                for col_idx, column in enumerate(table.columns, start=1):
                    key = column.key
                    value = row.values.get(key)

                    # Resolve formula keys to actual formulas
                    if isinstance(value, str) and value in formula_lib.formulas:
                        ws.cell(row=row_idx, column=col_idx, value=formula_lib.get(value))
                    elif column.formula:
                        # Column-level formula template
                        resolved = formula_lib.resolve(column.formula)
                        ws.cell(row=row_idx, column=col_idx, value=resolved)
                    else:
                        ws.cell(row=row_idx, column=col_idx, value=value)

    wb.save(output_path)


if __name__ == "__main__":
    output_file = Path(__file__).parent / "ServiceRequestWorkbook.xlsx"
    export_workbook_to_excel(service_request_workbook, formulas, output_file)
    print(f"Exported workbook to {output_file}")
